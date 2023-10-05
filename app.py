
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from openpyxl import workbook
from openpyxl import load_workbook
from time import sleep 


numero_oab = 133864

# Entrar no site consultado: https://pje-consulta-publica.tjmg.jus.br/
driver = webdriver.Chrome()
driver.get ('https://pje-consulta-publica.tjmg.jus.br/')
sleep(10)

# Inserir número da OAB e estado: //tag[atributo='valor]
campo_oab = driver.find_element(By.XPATH,"//input[@id='fPP:Decoration:numeroOAB']")
campo_oab.send_keys(numero_oab)
# Estado 
selecao_estado = driver.find_element(By.XPATH, "//select[@id='fPP:Decoration:estadoComboOAB']")
opcao_estado = Select(selecao_estado)
opcao_estado.select_by_visible_text('SP')

# Pesquisar
pesquisar = driver.find_element(By.XPATH, "//input[@id='fPP:searchProcessos']")
pesquisar.click()
sleep(10)

# Entrar em cada processo
processos = driver.find_elements(By.XPATH, "//b[@class='btn-block']")
for processo in processos:
    processo.click()
sleep(15)
# Ajustar tamanho da janela
janela = driver.window_handles
driver.switch_to.window(janela[-1])
driver.set_window_size(1920,1080)
# Coletar número dos processos
numero_processo = driver.find_elements(By.XPATH, "//div[@class='col-sm-12']")
numero_processo = numero_processo[0]
numero_processo = numero_processo.text
# Data
data_processo = driver.find_elements(By.XPATH, "//div[@class='col-sm-12']")
data_processo = data_processo[1]
data_processo = data_processo.text

# Extrair e armazenar as informações
movimentacoes = driver.find_elements(By.XPATH,"//div[@id='j_id132:processoEventoPanel_body']//tr[contains(@class,'rich-table-row')]//td//div//div//span")
lista_movimentacao = []
for movimentacao in movimentacoes:
    lista_movimentacao.append(movimentacao.text)

# Extrair para planilha
workbook = load_workbook('dados.xlsx')

try:
    # No caso de planilha existente:
    # Acessando a página do processo
    pagina_processo = workbook[numero_processo]

    # Criando o nome das colunas
    pagina_processo['A1'].value = "Número Processo"
    pagina_processo['B1'].value = "Data Distribuição"
    pagina_processo['C1'].value = "Movimentações"

    # Adicionando o número do processo
    pagina_processo['A2'].value = numero_processo

    # Adicionando data
    pagina_processo['B2'].value = data_processo

    # Adicionando movimentações
    for index, linha in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(lista_movimentacao), min_col=3, max_col=3)):
        for celula in linha:
            celula.value = lista_movimentacao[index]

    # Save the workbook after adding data to the existing sheet
    workbook.save('dados.xlsx')
    driver.close()
    sleep(5)
    driver.switch_to.window(driver.window_handles[0])

# Criando planilha do zero
except Exception as error:
    workbook.create_sheet(numero_processo)

    # Acessando a página do processo
    pagina_processo = workbook[numero_processo]

    # Nomeando as colunas
    pagina_processo['A1'].value = "Número Processo"
    pagina_processo['B1'].value = "Data Distribuição"
    pagina_processo['C1'].value = "Movimentações"

    # Adicionando numero dos processos
    pagina_processo['A2'].value = numero_processo

    # Adicionando data
    pagina_processo['B2'].value = data_processo

    # Adicionando movimentações
    for index, linha in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(lista_movimentacao), min_col=3, max_col=3)):
        for celula in linha:
            celula.value = lista_movimentacao[index]

    # Salvar os dados na planilha
    workbook.save('dados.xlsx')
    driver.close()
    sleep(5)
    driver.switch_to.window(driver.window_handles[0])


